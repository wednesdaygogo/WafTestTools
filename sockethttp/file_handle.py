import os
import random
import base64
import download_picture
from openpyxl import load_workbook

def get_filename(path):
    filelist = os.listdir(path)
    return filelist
def read_file(path):
    with open(path,encoding='utf-8',errors='ignore') as f:
        return f.read()

def write_file(path):
    count = 0
    filenames = get_filename(path)
    for filename in filenames:
        file_path = path + filename
        taget_filename = filename.replace(" ","")
        content = read_file(file_path)
        with open('./black_list/{}'.format(taget_filename),'w',encoding='utf-8') as f:
            f.write(content)
        print('写入成功 文件名:{}'.format(filename))
        count = count + 1
    print('共处理文件名{}个'.format(count))

def generate_random_str(ip,port,length=16,encode=None,filename="N"):#生成payload为随机字符串的http报文集合，可选择是否进行base64编码
    random_str = ''
    base_str = 'ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz1234567890`~!@#$%^&*()_-+=[{]};:\'\",<.>/?|\\'
    str_len = len(base_str) - 1
    for i in range(length):
        random_str += base_str[random.randint(0,str_len)]
    if encode =="base64":
        random_str = base64_encode(random_str)
        payload = generate_http_payload(ip, port, random_str,filename)
    else:
        payload = generate_http_payload(ip,port,random_str,filename)
    return payload

def generate_internet_str(ip,port,file,encode=None):#根据nft处理后的元数据生成现网的http报文集合，可选择是否进行base64编码
    content = read_file("./post_payload/{}".format(file))
    if encode == "base64":
        internet_str = base64_encode(content)
        payload = generate_http_payload(ip,port,internet_str,file)
    else:
        payload = generate_http_payload(ip, port, content, file)
    return payload

def generate_picture(ip,port,keyword=None,pages=10):#自动在百度图片中爬取指定关键字的图片并生成由图片组成的payload集合
    if os.path.exists("picture"):
        pass
    else:
        download_picture.get_images_from_baidu(keyword,pages,"picture")
        filenames = get_filename("picture")
        for filename in filenames:
            generate_http_payload(ip,port,read_file("picture\\"+filename))


def base64_encode(str):#工具函数，对指定串进行base64编码
    str_encode = base64.b64encode(str.encode('utf-8'))
    str_encode = str_encode.decode('utf-8')
    return str_encode

def generate_http_payload(ip,port,payload,filename):#将处理好的数据填写进指定的http模板中
    content = read_file('http.txt')
    if port == 80:
        host = ip
    else:
        host = ip+":"+str(port)
    http = content.replace("%ip",host).replace("%length",str(len(payload))).replace("%payload",payload).replace("%file",filename)
    return http

def whitefile_write(ip,port,type,length=16,num=2000,encode=None):#将生成好的http报文批量写入指定文件夹
    '''生成随机字符串文件可选择是否base64编码'''
    if type == "random":
        path = './white_list/simple_post/'
        path = path.strip()
        path = path.rstrip("\\")
        isexist = os.path.exists(path)
        if not isexist:
            os.makedirs(path)
        for i in range(0,num):
            with open(path+str(i)+'.txt','w',encoding='utf-8') as f:
                f.write(generate_random_str(ip,port,length,encode,str(i)))
                print("成功生成文件:{}".format(path+str(i)+'.txt'))
    #生成现网流量文件可选择是否base64编码
    elif type == "internet":
        path = './white_list/simple_post/'
        path = path.strip()
        path = path.rstrip("\\")
        isexist = os.path.exists(path)
        if not isexist:
            os.makedirs(path)
        dic = get_filename("./post_payload")
        for filename in dic:
            with open(path+filename,'w',encoding='utf-8') as f:
                f.write(generate_internet_str(ip,port,filename,encode))
                print("成功生成文件:{}".format(path+filename))
    #根据getclass.py生成的xlsx表格生成序列化测试数据
    elif type == "php_serilize":
        path = './black_list/php_serilize_payload/'
        path = path.strip()
        path = path.rstrip("\\")
        isexist = os.path.exists(path)
        if not isexist:
            os.makedirs(path)
        workbook_object = load_workbook(filename='phpggc_payload.xlsx')
        sheet = workbook_object.worksheets[0]
        max_row_num = sheet.max_row
        for i in range(2,max_row_num + 1):
            with open(path+sheet[i][0].value.replace('/','_') + '.txt','w',encoding='utf-8') as f:
                f.write(generate_http_payload(ip,port,sheet[i][1].value.strip(),sheet[i][0].value))
                print("成功生成文件:{}".format(path+sheet[i][0].value.replace('/','_') + '.txt'))
    elif type == "xxe":
        path = './black_list/aes_lanjie_payload/'
        path = path.strip()
        path = path.rstrip("\\")
        isexist = os.path.exists(path)
        if not isexist:
            os.makedirs(path)
        workbook_object = load_workbook(filename='lanjie.xlsx')
        sheet = workbook_object.worksheets[0]
        max_row_num = sheet.max_row
        for i in range(2,max_row_num + 1):
            with open(path+str(sheet[i][0].value) + '.txt','w',encoding='utf-8') as f:
                f.write(generate_http_payload(ip,port,sheet[i][1].value,str(sheet[i][0].value)))
                print("成功生成文件:{}".format(path+str(sheet[i][0].value) + '.txt'))
    elif type == "xlsx":
        path = './black_list/aes_lanjie_payload/'
        path = path.strip()
        path = path.rstrip("\\")
        isexist = os.path.exists(path)
        if not isexist:
            os.makedirs(path)
        workbook_object = load_workbook(filename='lanjie.xlsx')
        sheet = workbook_object.worksheets[0]
        max_row_num = sheet.max_row
        for i in range(2,max_row_num + 1):
            with open(path+str(sheet[i][0].value) + '.txt','w',encoding='utf-8') as f:
                f.write(sheet[i][1].value)
                print("成功生成文件:{}".format(path+str(sheet[i][0].value) + '.txt'))



if __name__ == '__main__':
    # 获取行数
    # workbook_object = load_workbook(filename='phpggc_payload.xlsx')
    # sheet = workbook_object.worksheets[0]
    # max_row_num = sheet.max_row
    # for i in range(2,max_row_num + 1):
    #     payload_dic = {'name':'','payload':''}
    #     payload_dic['name'] = sheet[i][0].value
    #     payload_dic['payload'] = sheet[i][1].value.strip()
    #     print(payload_dic)
    whitefile_write("99.99.99.88",80,"xlsx")

    # x= base64_encode("ss())(Lp;ph88;\\@#$RF>?\":L{PO\$)P{IPK\#\)J\)\$NMM\:Kls")
    # print(x)

    #generate_picture("99.99.99.88",80,"cat")
    #print(get_filename("white_payload"))
    #generate_random_str("99.99.99.88",80,1500,"base64")
    #generate_http_payload("99.99.99.88",800,"abc=name")
    #write_file('./black-list/')
    #whitefile_write(20000)
   # y = base64_encode('sss')
   # print(y)