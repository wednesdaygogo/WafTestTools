import socket
import os
import argparse
import sys
import time
from concurrent.futures import ThreadPoolExecutor,as_completed
import re
import requests
from openpyxl import load_workbook
import shutil

black_payload_path ={
    'php_code_injection':'./black_list/phpcodeinject_small_4k_payload/',
    'php_serilize':'./black_list/php_serilize_payload/',
    'xml_injection':'./black_list/xxe_payload/',
    'xray_php_serilize_payload':'./black_list/xray_php_serilize_payload/',
    'aes_payload':'./black_list/aes_lanjie_payload/'
}

white_payload_path = {
    'white_base64':'./white_list/white_payload_base64/',
    'white_picture':'./white_list/white_payload_picture/',
    'white_internet_post':'./white_list/venus_internet_post_payload/',
    'normal_string':'./white_list/normal_string_post_payload/',
}

def read_dir(path):
    filelist = []
    files = os.listdir(path)
    for file in files:
        filelist.append(path +file)
    return filelist
def read_file(path):
    with open(path,'r',encoding='utf-8') as f:
        content = f.read().replace('\r','').replace('\n','\r\n')
    return content+'\r\n\r\n'
def process_http_request(data):
    if data.startswith('GET') or data.startswith('POST') or data.startswith('PUT') or data.startswith('DELETE'):
        headers, body = data.split('\r\n\r\n', 1)
        headers = headers.split('\r\n')
        content_length_present = False

        for i, header in enumerate(headers):
            if header.startswith('Content-Length:'):
                headers[i] = 'Content-Length: ' + str(len(body))
                content_length_present = True
                break

        if not content_length_present:
            headers.append('Content-Length: ' + str(len(body)))

        return '\r\n'.join(headers) + '\r\n\r\n' + body
    else:
        return data


def copy_file(src,dst):
    isexist = os.path.exists(dst)
    if not isexist:
        os.makedirs(dst)
    shutil.copy(src,dst)
def send(ip,port,content):
    #print(content)
    url = ip
    port = port
    socket.setdefaulttimeout(5)
    sock = socket.socket(socket.AF_INET,socket.SOCK_STREAM)
    sock.connect((url,port))
    sock.sendall(content)
    #response = b''
    res = sock.recv(1024)
    # while rec:
    #     response += rec
    #     rec = sock.recv(1024)
    result = res.decode(errors='ignore')
    sock.close()
    return result

def main(url,port,content,file):
    try:
        result  = send(url,port,content)
        #print(result)
    except TimeoutError:
        print(file + ' 连接超时')
        return 2
        #copy_file(file,'.\\outtime\\')
    except ConnectionResetError:
        print(file + ' 已拦截')
        return 1
    except ConnectionAbortedError:
        #print(result)
        if result.find('403') == -1:
            print(file + ' 未拦截')
            return 0
            #copy_file(file,'.\\QMNO\\')
        else:
            print(file + ' 已拦截')
            return 1
            #copy_file(file,'.\\command-wubao\\')


def muti_thread_test(url,port,dir,threads):
    success = 0
    faild = 0
    out_time = 0
    future_list = []
    files = read_dir(dir)
    with ThreadPoolExecutor(threads) as executor:
        for file in files:
            content = read_file(file)
            a = [url,port,content,file]
            future_list.append(executor.submit(lambda p: main(*p),a))
    for future in as_completed(future_list):
        results = future.result()
        print(results)
        if results == 1:
            success = success +1
        elif results == 0:
            faild = faild + 1
        else:
            out_time = out_time + 1
    print("总共检测payload："+str(success+faild+out_time))
    print("拦截数量：" + str(success))
    print("未拦截数量：" + str(faild))
    print("超时数量：" + str(out_time))
    print("检测率："+str(success/success+faild+out_time))

def send_from_xlsx_plus(url,port,path,path_2,column1,column2=None):

    if os.path.splitext(path)[-1] != '.xlsx':
        print('-x 参数需为xlsx后缀文件')
        sys.exit()
    success = 0
    count = 0
    workbook_object = load_workbook(filename=path)
    sheet = workbook_object.worksheets[0]
    print(sheet.title)
    max_row_num = sheet.max_row
    print(max_row_num)

    for i in range(2, max_row_num + 1):
        try:
            if column2 == None:
                content = str(sheet[i][column1].value).lstrip().replace("\n","\r\n")+"\r\n\r\n"
                content = process_http_request(content)
                # content = str(sheet[i][column1].value).strip('"').replace("\\r\\n", "\r\n").replace('\\"', '"')
            else:
                content1 = str(sheet[i][column1].value).strip('"').replace("\\r\\n","\r\n").replace('\\"','"')
                content2 = str(sheet[i][column2].value).strip('"').replace("\\r\\n","\r\n").replace('\\"','"') + "\r\n\r\n"
                content1 = re.sub(r'Content-Length:[\s0-9]*','Content-Length: {}\r\n'.format(len(content2)),content1)
                content = content1 + content2
            #print(content)
            content = content.encode("utf-8")
        except AttributeError:
            continue
        try:
            count = count + 1
            result = send(url, int(port), content)
        except TimeoutError:
            print('序号:' + str(count) + ' ' + 'line: '+ str(i) + ' 连接超时')
            # copy_file(file,'.\\outtime\\')
            continue
        except ConnectionResetError:
            time.sleep(0.6)
            event_name = check_eventname_form_waf()
            print('序号:' + str(count) + ' ' + 'line: '+ str(i) + ' 已拦截  ' + '上报事件:{}'.format(event_name))
            cell_event_name = sheet.cell(i,1)
            cell_event_name.value = event_name
            cell_event_action = sheet.cell(i,2)
            cell_event_action.value = '拦截'
            success = success + 1
            continue
        except ConnectionAbortedError:
            time.sleep(0.6)
            event_name = check_eventname_form_waf()
            print('序号:' + str(count) + ' ' + 'line: ' + str(i) + ' 已拦截  ' + '上报事件:{}'.format(event_name))
            cell_event_name = sheet.cell(i, 1)
            cell_event_name.value = event_name
            cell_event_action = sheet.cell(i, 2)
            cell_event_action.value = '拦截'
            success = success + 1
            continue
        except:
            print('序号:' + str(count) + ' ' + 'line: ' + str(i) + ' 未拦截')
            cell_event_action = sheet.cell(i, 2)
            cell_event_action.value = '未拦截'
            continue
        # print(result)
        if result.find('403') == -1 and result.find('307') == -1:
            print('序号:' + str(count) + ' ' + 'line: '+ str(i) + ' 未拦截')
            cell_event_action = sheet.cell(i, 2)
            cell_event_action.value = '未拦截'
            # copy_file(file,'.\\aes_bypass\\')
        else:
            time.sleep(0.6)
            event_name = check_eventname_form_waf()
            print('序号:' + str(count) + ' ' + 'line: '+ str(i) + ' 已拦截  ' + '上报事件:{}'.format(event_name) )
            cell_event_name = sheet.cell(i, 1)
            cell_event_name.value = event_name
            cell_event_action = sheet.cell(i, 2)
            cell_event_action.value = '拦截'
            success = success + 1
            # copy_file(file,'.\\command-wubao\\')
    workbook_object.save(path_2)
    print("一共发送样本数量：{}".format(max_row_num - 1))
    print("拦截数：{}".format(success))
    print("未拦截数：{}".format(max_row_num - success - 1))
    print("检测率：" + str(success / (max_row_num - 1)))

def check_eventname_form_waf(open=1):
    if open == 1:
        return "事件名称查询功能未开启"
    header = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/78.0.3904.108 Safari/537.36',
        'Cookie': 'SID=s%3A0SUqI9px8ef40VTt9ny6m3mSmyUxhqiM.msnruVtgrlAx38bubFluco7wC9ieEwwHsuthi2oaaxE; VWPHPUSERID=adm',
        'Content-Type':'application/json',
        'Connection':'close',
        'Authorization': 'Bearer eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJ1c2VybmFtZSI6ImFkbSIsImtleSI6IjE2OTEzNzUzMzk4MDEtMC43NTkzMjQ0MDY4MzgwMDMiLCJpYXQiOjE2OTEzODg1NzQsImV4cCI6MTY5MTQxNzM3NH0.EXnZCwj7YpLsOqBK_vxwktlALLsHtLIfLcvVd3jFw0U'
    }
    data = {"page":1,"limit":1,"isEnglish":0,"searchType":"ad","searchstr":"","auditlinkage_dstip":"","auditlinkage_uv":"","auditlinkage_pv":"","auditlinkage_port":0,"auditlinkage_wafip":"10.51.15.186","filters":[],"chk_time":"on","chk_evt_name":"on","chk_evt_group":"on","chk_evt_level":"on","chk_x_forwarded_for":"on","chk_srcip_str":"on","chk_srcport":"on","chk_dstip_str":128,"chk_dstport":"on","chk_action":"on","chk_rawguid":"on"}
    url = "http://10.51.15.186/securityeventmonitoring/eventmonlist"
    # response = requests.post(url=url, json=data,headers=header)
    # result = response.json()
    # return result['data']['rows'][0]['evt_name']
    while True:
        response_1 = requests.post(url=url, json=data,headers=header)
        response_2 = requests.post(url=url, json=data, headers=header)
        result_1 = response_1.json()
        result_2 = response_2.json()
        if result_1['data']['rows'][0]['evt_name'] == result_2['data']['rows'][0]['evt_name']:
            break

    return result_2['data']['rows'][0]['evt_name']

def single_thread_test(url,port,file_dir):
    dir = []
    success = 0
    count = 0
    if file_dir == 'white_all':
        for key,value in white_payload_path.items():
            dir += read_dir(value)
    elif file_dir == 'black_all':
        for key,value in black_payload_path.items():
            dir += read_dir(value)
    else:
        if file_dir[-1] != '\\' and file_dir[-1] != '/':
            file_dir = file_dir + '/'
        dir = read_dir(file_dir)
    total = len(dir)
    #print(total)
    for file in dir:
        content = read_file(file)
        #print(content)
        try:
            count = count + 1
            #print(process_http_request(content))
            result = send(url,int(port),process_http_request(content).encode('utf-8'))
            #print(process_http_request(content))
            #print(result)
        except TimeoutError:
            print('序号:'+str(count) + ' ' + file + ' 连接超时')
            #copy_file(file,'.\\qm_yes\\')
            continue
        except ConnectionResetError:
            print('序号:'+str(count) + ' ' + file + ' 已拦截')
            success = success + 1
            #copy_file(file, '.\\qm_yes\\')
            continue
        except ConnectionAbortedError:
            print('序号:' + str(count) + ' ' + file + '已拦截')
            success = success + 1
            # copy_file(file, '.\\qm_yes\\')
            continue
        except:
            print('序号:' + str(count) + ' ' + file + ' 未拦截')
            continue
            copy_file(file, './pycat_no/')
        #print(result)
        if result.find('403') == -1:
            print('序号:'+str(count) + ' ' + file + ' 未拦截')
            copy_file(file, './pycat_no/')
        else:
            print('序号:'+str(count) + ' ' + file + ' 已拦截')
            success = success + 1
            copy_file(file,'./php_wubao/')
    print("一共发送样本数量：{}".format(total))
    print("拦截数：{}".format(success))
    print("未拦截数：{}".format(total-success))
    print("检测率："+str(success/total))



if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument("-u", "--url", help="Target ip;Example:10.88.1.220")
    parser.add_argument("-p", "--port", help="Target port;Example:8001")
    parser.add_argument("-d", "--dir", help="Target file;Example:D:\pythonProject\sockethttp\\text")
    parser.add_argument("-t", "--threads", help="thread number;Example:-t 10代表10个线程")
    parser.add_argument("-w","--white",help="白流量测试；Example: -w white_base64 表示测试white_base64类型白流量，all测试全部类型")
    parser.add_argument("-b", "--black", help="黑流量测试；Example: -b php_serilize 表示测试php_serilize类型黑流量，all测试全部类型")
    parser.add_argument("-xp", "--xlsxplus", help="从xlsx中读取payload测试，用于测试请求体和请求头分别在两列的表，测试完成后在原表的1，2列添加事件名称和测试结果")
    parser.add_argument("-df", "--dstfile", help="xlsxplus测试功能结果生成的目标文件")
    parser.add_argument("-c1", "--column1", help="xlsxplus测试功能参与拼接的列数1(有先后顺序，c1拼接在c2之前,c2不写的话就只读c1)")
    parser.add_argument("-c2", "--column2", help="xlsxplus测试功能参与拼接的列数2(有先后顺序，c1拼接在c2之前，c2不写的话就只读c1)")
    args = parser.parse_args()
    if args.threads == None:
        if args.white != None:
            if args.white == 'all':
                single_thread_test(args.url,int(args.port),'white_all')
            else:
                single_thread_test(args.url,int(args.port),white_payload_path[args.white])
        if args.black != None:
            if args.black == 'all':
                single_thread_test(args.url,int(args.port),'black_all')
            else:
                single_thread_test(args.url,int(args.port),black_payload_path[args.black])
        if args.dir != None:
            single_thread_test(args.url, int(args.port),args.dir)
        if args.xlsxplus != None:
            if args.column2 == None:
                send_from_xlsx_plus(args.url,int(args.port), args.xlsxplus,args.dstfile, int(args.column1))
            else:
                send_from_xlsx_plus(args.url, int(args.port), args.xlsxplus, args.dstfile, int(args.column1),int(args.column2))

    else:
        muti_thread_test(args.url,int(args.port),args.dir,int(args.threads))

    # txt = read_file('D:\pythonProject\sockethttp\\test1\\16.txt')
    # print(process_http_request(txt))

