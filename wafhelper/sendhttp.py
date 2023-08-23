import os
import re
import sys
from filehandle import *
from openpyxl import load_workbook
import HackRequests

def send(ip,port,content):
    hack = HackRequests.hackRequests()
    real_host = ip +":"+port
    response = hack.httpraw(raw=content,real_host=real_host)


def send_from_xlsx(url,port,path,path_2,column1,column2=None):
    success = 0
    count = 0
    if os.path.splitext(path)[-1] != '.xlsx':
        print('-x 参数需为xlsx后缀文件')
        sys.exit()
    workbook_object = load_workbook(filename=path)
    sheet = workbook_object.worksheets[0]
    max_row_num = sheet.max_row
    print(max_row_num)

    for i in range(2, max_row_num + 1):
        try:
            if column2 == None:
                content = str(sheet[i][column1].value) + "\r\n"
                # content = str(sheet[i][column1].value).strip('"').replace("\\r\\n", "\r\n").replace('\\"', '"')
            else:
                content1 = str(sheet[i][column1].value).strip('"').replace("\\r\\n", "\r\n").replace('\\"', '"')
                content2 = str(sheet[i][column2].value).strip('"').replace("\\r\\n", "\r\n").replace('\\"','"') + "\r\n\r\n"
                content1 = re.sub(r'Content-Length:[\s0-9]*', 'Content-Length: {}\r\n'.format(len(content2)), content1)
                content = content1 + content2
            # print(content)
            content = content.encode("utf-8")
        except AttributeError:
            continue
        try:
            count = count + 1
            result = send(url, port, content)

            # print(result)
        except TimeoutError:
            print('序号:' + str(count) + ' ' + 'line: ' + str(i) + ' 连接超时')
            # copy_file(file,'.\\outtime\\')
            continue
        except ConnectionResetError:
            # time.sleep(0.4)
            event_name = check_eventname_form_waf()
            print('序号:' + str(count) + ' ' + 'line: ' + str(i) + ' 已拦截  ' + '上报事件:{}'.format(event_name))
            cell_event_name = sheet.cell(i, 1)
            cell_event_name.value = event_name
            cell_event_action = sheet.cell(i, 2)
            cell_event_action.value = '拦截'
            success = success + 1
            continue
        except ConnectionAbortedError:
            continue
        except:
            print('序号:' + str(count) + ' ' + 'line: ' + str(i) + ' 未拦截')
            cell_event_action = sheet.cell(i, 2)
            cell_event_action.value = '未拦截'
            continue
        # print(result)
        if result.find('403') == -1:
            print('序号:' + str(count) + ' ' + 'line: ' + str(i) + ' 未拦截')
            cell_event_action = sheet.cell(i, 2)
            cell_event_action.value = '未拦截'
            # copy_file(file,'.\\aes_bypass\\')
        else:
            # time.sleep(0.4)
            event_name = check_eventname_form_waf()
            print('序号:' + str(count) + ' ' + 'line: ' + str(i) + ' 已拦截  ' + '上报事件:{}'.format(event_name))
            cell_event_name = sheet.cell(i, 1)
            cell_event_name.value = event_name
            cell_event_action = sheet.cell(i, 2)
            cell_event_action.value = '拦截'
            success = success + 1
            # copy_file(file,'.\\command-wubao\\')
    workbook_object.save(path_2)
    print("一共发送样本数量：{}".format(max_row_num - 1))
    print("拦截数：{}".format(success))
    print("未拦截数：{}".format(max_row_num - success - 1))
    print("检测率：" + str(success / max_row_num))

def send_from_pcap():
    pass
def send_from_dic():
    pass
def send_from_txt():
    pass